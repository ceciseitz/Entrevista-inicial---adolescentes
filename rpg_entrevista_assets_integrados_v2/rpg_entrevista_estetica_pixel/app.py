
from flask import Flask, render_template, request, jsonify, send_file
from pathlib import Path
from datetime import datetime
from openpyxl import Workbook, load_workbook

app = Flask(__name__)
DATA_DIR = Path("data")
DATA_DIR.mkdir(exist_ok=True)
EXCEL_PATH = DATA_DIR / "entrevistas_rpg.xlsx"

HEADERS = [
    "fecha", "nombre", "pronombres", "edad", "mood_inicial",
    "fortalezas", "dificultades", "personas_importantes", "lo_que_me_mueve",
    "autoestima", "energia", "concentracion", "animo", "ansiedad", "sueno", "felicidad", "enojo",
    "casa_mejor", "casa_dificil", "casa_siento", "casa_cambio",
    "escuela_mejor", "escuela_dificil", "escuela_siento", "escuela_cambio",
    "amistades_mejor", "amistades_dificil", "amistades_siento", "amistades_cambio",
    "internet_mejor", "internet_dificil", "internet_siento", "internet_cambio",
    "cabeza_mejor", "cabeza_dificil", "cabeza_siento", "cabeza_cambio",
    "cuerpo_mejor", "cuerpo_dificil", "cuerpo_siento", "cuerpo_cambio",
    "recarga", "drena",
    "mito_demasiado", "mito_empezar_tareas", "mito_pensar_hablar", "mito_distraigo",
    "mito_comparo", "mito_no_encajo", "mito_cumplidos", "mito_lograr", "mito_frustro", "mito_autocritica",
    "mision_confianza", "mision_concentracion", "mision_organizacion", "mision_sueno",
    "mision_relaciones", "mision_emociones", "mision_autotrato", "mision_escuela",
    "mision_especial"
]

def ensure_workbook():
    if not EXCEL_PATH.exists():
        wb = Workbook()
        ws = wb.active
        ws.title = "Entrevistas"
        ws.append(HEADERS)
        wb.save(EXCEL_PATH)

@app.route("/")
def index():
    return render_template("index.html")

@app.route("/save", methods=["POST"])
def save():
    ensure_workbook()
    payload = request.get_json(force=True) or {}
    wb = load_workbook(EXCEL_PATH)
    ws = wb.active
    row = []
    for h in HEADERS:
        if h == "fecha":
            row.append(datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
        else:
            value = payload.get(h, "")
            if isinstance(value, list):
                value = "; ".join(value)
            row.append(value)
    ws.append(row)
    wb.save(EXCEL_PATH)
    return jsonify({"ok": True, "message": "Entrevista guardada", "excel": str(EXCEL_PATH)})

@app.route("/download")
def download():
    ensure_workbook()
    return send_file(EXCEL_PATH, as_attachment=True)

if __name__ == "__main__":
    app.run(debug=True)
